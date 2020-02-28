import random
from openpyxl import load_workbook

workbook = load_workbook('club_sorter.xlsx')
import_data = workbook['Input Data']
sorted_data = workbook['Sorted Data']

# variable declaration
CLUB_SPACES = 20
CURRENT_ROW = 2
STUDENT_COUNT = 315
A_CLUBS_COUNT = 29
B_CLUBS_COUNT = 28
super_yay = 0


class Clubs:
    def __init__(self, name, spots_available, club_desire, club_desire_total, assigned_students, assigned_students_freshmen,
                 assigned_students_sophomore, assigned_students_junior, assigned_students_senior,
                 assigned_students_total):
        self.name = name
        self.spots_available = spots_available
        self.club_desire = club_desire
        self.club_desire_total = club_desire_total
        self.assigned_students = assigned_students
        self.assigned_students_freshmen = assigned_students_freshmen
        self.assigned_students_sophomore = assigned_students_sophomore
        self.assigned_students_junior = assigned_students_junior
        self.assigned_students_senior = assigned_students_senior
        self.assigned_students_total = assigned_students_total

class Student:
    def __init__(self, email, name, grade, choice_1, choice_2, choice_3, choice_4, choice_5, choice_6, assigned_a, assigned_b):
        self.email = email
        self.name = name
        self.grade = grade
        self.choice_1 = choice_1
        self.choice_2 = choice_2
        self.choice_3 = choice_3
        self.choice_4 = choice_4
        self.choice_5 = choice_5
        self.choice_6 = choice_6
        self.assigned_a = assigned_a
        self.assigned_b = assigned_b


student_list = list()
student_row = CURRENT_ROW
while student_row <= STUDENT_COUNT:
    student = Student(import_data.cell(student_row, 2).value, import_data.cell(student_row, 3).value,
                      import_data.cell(student_row, 4).value, import_data.cell(student_row, 5).value,
                      import_data.cell(student_row, 6).value, import_data.cell(student_row, 7).value,
                      import_data.cell(student_row, 8).value, import_data.cell(student_row, 9).value,
                      import_data.cell(student_row, 10).value, None, None)
    student_list.append(student)
    student_row += 1

club_list_a = list()
club_row_a = CURRENT_ROW
while club_row_a <= A_CLUBS_COUNT:
    club_a = Clubs(import_data.cell(club_row_a, 13).value, CLUB_SPACES, list(), 0, list(), 0, 0, 0, 0, 0)
    club_list_a.append(club_a)
    club_row_a += 1

club_list_b = list()
club_row_b = CURRENT_ROW
while club_row_b <= B_CLUBS_COUNT:
    club_b = Clubs(import_data.cell(club_row_b, 14).value, CLUB_SPACES, list(), 0, list(), 0, 0, 0, 0, 0)
    club_list_b.append(club_b)
    club_row_b += 1


def grade_total():
    if student.grade == 'Freshman':
        club.assigned_students_freshmen += 1

    elif student.grade == 'Sophomore':
        club.assigned_students_sophomore += 1

    elif student.grade == 'Junior':
        club.assigned_students_junior += 1

    elif student.grade == 'Senior':
        club.assigned_students_senior += 1


club_names_a = list()
for club in club_list_a:
    club_names_a.append(club.name)

club_names_b = list()
for club in club_list_b:
    club_names_b.append(club.name)

# first iteration
for club in club_list_a:
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_1 == club.name and student.assigned_a is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, (club.club_desire_total + club.assigned_students_total) - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        club_names_b = list()
        for club in club_list_b:
            club_names_b.append(club.name)

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if club.name not in club_list_b:
                        if student.email not in club.assigned_students:
                            if student.choice_1 == club.name and student.assigned_a is None:
                                club.assigned_students.append(student.email)
                                student.assigned_a = student.choice_1
                                club.assigned_students_total += 1
                                grade_total()

                        elif club.name in club_list_b:
                            if student.choice_2 in club_list_a:
                                if student.email not in club.assigned_students:
                                    if student.choice_1 == club.name and student.assigned_b is None:
                                        club.assigned_students.append(student.email)
                                        student.assigned_b = student.choice_1
                                        club.assigned_students_total += 1
                                        grade_total()

                            elif student.choice_2 in club_list_b:
                                if student.email not in club.assigned_students:
                                    if student.choice_1 == club.name and student.assigned_a is None:
                                        club.assigned_students.append(student.email)
                                        student.assigned_a = student.choice_1
                                        club.assigned_students_total += 1
                                        grade_total()

    # Run this chunk if the two totals are less than 20
    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if club.name not in club_names_b:
                if student.choice_1 == club.name and student.assigned_a is None:
                    if student.email not in club.assigned_students:
                        club.assigned_students.append(student.email)
                        student.assigned_a = student.choice_1
                        club.assigned_students_total += 1
                        grade_total()

            elif club.name in club_names_b:
                if student.choice_2 in club_names_a:
                    if student.email not in club.assigned_students:
                        if student.choice_1 == club.name and student.assigned_b is None:
                            club.assigned_students.append(student.email)
                            student.assigned_b = student.choice_1
                            club.assigned_students_total += 1
                            grade_total()

                elif student.choice_2 in club_list_b:
                    if student.email not in club.assigned_students:
                        if student.choice_1 == club.name and student.assigned_a is None:
                            club.assigned_students.append(student.email)
                            student.assigned_a = student.choice_1
                            club.assigned_students_total += 1
                            grade_total()

for club in club_list_b:
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))

    for student in student_list:
        if student.choice_1 == club.name and student.assigned_b is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, (club.club_desire_total + club.assigned_students_total) - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.email not in club.assigned_students:
                        club.assigned_students.append(student.email)
                        student.assigned_b = student.choice_1
                        club.assigned_students_total += 1
                        grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_1 == club.name and student.assigned_b is None:
                if student.email not in club.assigned_students:
                    if student.assigned_a != club.name:
                        club.assigned_students.append(student.email)
                        student.assigned_b = student.choice_1
                        club.assigned_students_total += 1
                        grade_total()

# second iteration
for club in club_list_a:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_2 == club.name and student.assigned_a is None and student.choice_1 != club.name:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if club.name not in club_list_b:
                        if student.choice_2 == club.name and student.assigned_a is None:
                            if student.email not in club.assigned_students:
                                club.assigned_students.append(student.email)
                                student.assigned_a = student.choice_2
                                club.assigned_students_total += 1
                                grade_total()

                    elif club.name in club_list_b:
                        if student.choice_1 in club_list_a and student.assigned_b is None:
                            if student.email not in club.assigned_students:
                                club.assigned_students.append(student.email)
                                student.assigned_a = student.choice_2
                                club.assigned_students_total += 1
                                grade_total()

                        elif student.choice_1 in club_list_b and student.assigned_a is None:
                            if student.email not in club.assigned_students:
                                club.assigned_students.append(student.email)
                                student.assigned_a = student.choice_2
                                club.assigned_students_total += 1
                                grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_2 == club.name and student.assigned_a is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_2
                    club.assigned_students_total += 1
                    grade_total()

for club in club_list_b:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_2 == club.name and student.assigned_b is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.choice_2 == club.name and student.assigned_b is None:
                        if student.email not in club.assigned_students:
                            if student.assigned_a != club.name:
                                club.assigned_students.append(student.email)
                                student.assigned_b = student.choice_2
                                club.assigned_students_total += 1
                                grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_2 == club.name and student.assigned_b is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_2
                    club.assigned_students_total += 1
                    grade_total()


# third iteration
for club in club_list_a:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_3 == club.name and student.assigned_a is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.choice_3 == club.name and student.assigned_a is None:
                        if student.email not in club.assigned_students:
                            club.assigned_students.append(student.email)
                            student.assigned_a = student.choice_3
                            club.assigned_students_total += 1
                            grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_3 == club.name and student.assigned_a is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_3
                    club.assigned_students_total += 1
                    grade_total()

for club in club_list_b:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_3 == club.name and student.assigned_b is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.choice_3 == club.name and student.assigned_b is None:
                        if student.email not in club.assigned_students:
                            if student.assigned_a != club.name:
                                club.assigned_students.append(student.email)
                                student.assigned_b = student.choice_3
                                club.assigned_students_total += 1
                                grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_3 == club.name and student.assigned_b is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_3
                    club.assigned_students_total += 1
                    grade_total()

# fourth iteration
for club in club_list_a:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_4 == club.name and student.assigned_a is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.choice_4 == club.name and student.assigned_a is None:
                        if student.email not in club.assigned_students:
                            club.assigned_students.append(student.email)
                            student.assigned_a = student.choice_4
                            club.assigned_students_total += 1
                            grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_4 == club.name and student.assigned_a is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_4
                    club.assigned_students_total += 1
                    grade_total()

for club in club_list_b:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_4 == club.name and student.assigned_b is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.choice_4 == club.name and student.assigned_b is None:
                        if student.email not in club.assigned_students:
                            if student.assigned_a != club.name:
                                club.assigned_students.append(student.email)
                                student.assigned_b = student.choice_4
                                club.assigned_students_total += 1
                                grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_4 == club.name and student.assigned_b is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_4
                    club.assigned_students_total += 1
                    grade_total()


# fifth iteration
for club in club_list_a:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_5 == club.name and student.assigned_a is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.choice_5 == club.name and student.assigned_a is None:
                        if student.email not in club.assigned_students:
                            club.assigned_students.append(student.email)
                            student.assigned_a = student.choice_5
                            club.assigned_students_total += 1
                            grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_5 == club.name and student.assigned_a is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_5
                    club.assigned_students_total += 1
                    grade_total()

for club in club_list_b:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_5 == club.name and student.assigned_b is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.choice_5 == club.name and student.assigned_b is None:
                        if student.email not in club.assigned_students:
                            if student.assigned_a != club.name:
                                club.assigned_students.append(student.email)
                                student.assigned_b = student.choice_5
                                club.assigned_students_total += 1
                                grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_5 == club.name and student.assigned_b is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_5
                    club.assigned_students_total += 1
                    grade_total()

# sixth iteration
for club in club_list_a:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_6 == club.name and student.assigned_a is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.choice_6 == club.name and student.assigned_a is None:
                        if student.email not in club.assigned_students:
                            club.assigned_students.append(student.email)
                            student.assigned_a = student.choice_6
                            club.assigned_students_total += 1
                            grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_6 == club.name and student.assigned_a is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_6
                    club.assigned_students_total += 1
                    grade_total()

for club in club_list_b:
    club.club_desire_total = 0
    club.club_desire = list()
    if str(club.name) not in club.assigned_students:
        club.assigned_students.append(str(club.name))
    # New
    for student in student_list:
        if student.choice_6 == club.name and student.assigned_b is None:
            club.club_desire.append(student.email)
            club.club_desire_total += 1

    chosen_indexes = list()

    if club.club_desire_total + club.assigned_students_total > 20:
        spots_left = 20 - club.assigned_students_total
        for random_run in range(spots_left):
            while True:
                chosen_index = random.randint(0, club.club_desire_total - 1)
                if chosen_index not in chosen_indexes:
                    chosen_indexes.append(chosen_index)
                    break
                else:
                    continue

        for index in chosen_indexes:
            for student in student_list:
                if student.email == club.club_desire[index]:
                    if student.choice_6 == club.name and student.assigned_b is None:
                        if student.email not in club.assigned_students:
                            if student.assigned_a != club.name:
                                club.assigned_students.append(student.email)
                                student.assigned_b = student.choice_6
                                club.assigned_students_total += 1
                                grade_total()

    elif club.club_desire_total + club.assigned_students_total < 20:
        for student in student_list:
            if student.choice_6 == club.name and student.assigned_b is None:
                if student.email not in club.assigned_students:
                    club.assigned_students.append(student.email)
                    student.assigned_a = student.choice_6
                    club.assigned_students_total += 1
                    grade_total()

for student in student_list:
    for club in club_list_a:
        if student.email in club.assigned_students:
            student.assigned_a = club.name

    for club in club_list_b:
        if student.email in club.assigned_students:
            student.assigned_b = club.name


for student in student_list:
    if student.assigned_a is None:
        for club in club_list_a:
            if club.assigned_students[0] == 'STUDY HALL':
                club.assigned_students.append(student.email)
                student.assigned_a = 'STUDY HALL'

    elif student.assigned_b is None:
        for club in club_list_b:
            if club.assigned_students[0] == 'STUDY HALL':
                club.assigned_students.append(student.email)
                student.assigned_b = 'STUDY HALL'

for student in student_list:
    if student.assigned_a == student.choice_1 and student.assigned_b == student.choice_2:
        super_yay += 1

    elif student.assigned_a == student.choice_2 and student.assigned_b == student.choice_1:
        super_yay += 1

study_hall_students = list()

for club in club_list_a:
    if club.name == 'STUDY HALL':
        for student in club.assigned_students:
            study_hall_students.append(student)

for club in club_list_a:
    if club.name == 'STUDY HALL':
        for student in club.assigned_students:
            study_hall_students.append(student)

a_week_hall = round(len(study_hall_students) / 2)

value_row = 1
for club in club_list_a:
    list_position = 1
    for name in club.assigned_students:
        sorted_data.cell(value_row, list_position, name).value
        list_position += 1
    value_row += 1

value_row = 1 + A_CLUBS_COUNT
for club in club_list_b:
    list_position = 1
    for name in club.assigned_students:
        sorted_data.cell(value_row, list_position, name).value
        list_position += 1
    value_row += 1

for club in club_list_a:
    print(club.assigned_students_total, 'entries', club.assigned_students_freshmen, club.assigned_students_sophomore,
          club.assigned_students_junior, club.assigned_students_senior, club.assigned_students)
for club in club_list_b:
    print(club.assigned_students_total, 'entries', club.assigned_students_freshmen, club.assigned_students_sophomore,
          club.assigned_students_junior, club.assigned_students_senior, club.assigned_students)

print(round(super_yay/STUDENT_COUNT * 100, 2), "%")

workbook.save('club_sorter.xlsx')
